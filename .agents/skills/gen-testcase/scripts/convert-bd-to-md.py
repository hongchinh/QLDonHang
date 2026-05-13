#!/usr/bin/env python3
"""
convert-bd-to-md.py — BD source (.xlsx) → BD.md + images

Self-contained script (no external imports from scripts/).
Dependencies: openpyxl (see requirements.txt).

Usage:
    python3 convert-bd-to-md.py --input <source.xlsx> --output <BD.md> --images-dir <dir>
"""

import argparse
import os
import shutil
import sys
import zipfile
from pathlib import Path


# ---------------------------------------------------------------------------
# Global settings loader
# ---------------------------------------------------------------------------

_SETTINGS_PATH = os.path.join(os.path.dirname(__file__), "..", "settings.yaml")

def _load_settings():
    """Load skill-level settings.yaml. Prints warning and returns empty dict on failure."""
    try:
        import yaml
    except ImportError:
        print("WARNING: pyyaml not installed — settings.yaml not loaded, using empty defaults.", file=sys.stderr)
        return {}
    if not os.path.exists(_SETTINGS_PATH):
        print(f"WARNING: settings.yaml not found at {_SETTINGS_PATH} — using empty defaults.", file=sys.stderr)
        return {}
    with open(_SETTINGS_PATH, "r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}

_SETTINGS = _load_settings()


def _get(path, default=None):
    """Get a nested value from settings by dot-separated path."""
    node = _SETTINGS
    for key in path.split("."):
        if isinstance(node, dict):
            node = node.get(key)
        else:
            return default
    return node if node is not None else default


KV_MAX_COLS = _get("bd_conversion.key_value_max_cols", 3)
DT_MIN_COLS = _get("bd_conversion.data_table_min_cols", 4)
MIN_ROWS = _get("bd_conversion.min_rows_for_classification", 2)
KV_THRESHOLD = _get("bd_conversion.key_value_qualifying_threshold", 0.6)
LABEL_MAX_CHARS = _get("bd_conversion.field_label_max_chars", 40)


# ---------------------------------------------------------------------------
# Render helpers (inlined from old xlsx-render-helpers.py)
# ---------------------------------------------------------------------------

def format_cell_value(val):
    """Return str(val).strip() or empty string."""
    if val is None:
        return ""
    return str(val).strip()


def _active_column_indices(rows_data, max_col):
    """Return sorted list of 0-based column indices that have any non-empty cell."""
    active = set()
    for row in rows_data:
        for ci in range(max_col):
            if ci < len(row) and row[ci]:
                active.add(ci)
    return sorted(active)


def detect_section_type(rows_data):
    """
    Classify a section as 'key_value', 'data_table', or 'plain'.

    key_value:
      - max used columns <= KV_MAX_COLS
      - >= MIN_ROWS rows
      - >= KV_THRESHOLD of rows have a first-cell label that is non-empty, < LABEL_MAX_CHARS chars, no newline
    data_table:
      - max used columns >= DT_MIN_COLS
      - >= MIN_ROWS rows
    Otherwise: plain
    """
    if not rows_data or len(rows_data) < MIN_ROWS:
        return "plain"

    max_col = max((len(r) for r in rows_data), default=0)
    active_cols = _active_column_indices(rows_data, max_col)
    used_cols = len(active_cols)

    if used_cols >= DT_MIN_COLS:
        return "data_table"

    if used_cols <= KV_MAX_COLS:
        qualifying = 0
        for row in rows_data:
            first = row[0] if row else ""
            if first and len(first) < LABEL_MAX_CHARS and "\n" not in first:
                qualifying += 1
        if qualifying / len(rows_data) >= KV_THRESHOLD:
            return "key_value"

    return "plain"


def render_key_value_section(rows, max_col):
    """
    Render as | Key | Value | table.
    Value = remaining non-empty cells joined with ' | '.
    Escape | as \\|, newlines as <br>.
    """
    lines = ["| Key | Value |", "| --- | --- |"]
    for row in rows:
        key = row[0] if row else ""
        key = key.replace("|", r"\|").replace("\n", "<br>")
        value_parts = []
        for ci in range(1, max_col):
            cell = row[ci] if ci < len(row) else ""
            if cell:
                value_parts.append(cell.replace("|", r"\|").replace("\n", "<br>"))
        value = " | ".join(value_parts)
        lines.append(f"| {key} | {value} |")
    return "\n".join(lines)


def render_data_table(rows, max_col):
    """
    Render as markdown table. First row is header.
    Strip all-empty columns. Empty header cells named Col{1-based}.
    """
    if not rows:
        return ""

    active = _active_column_indices(rows, max_col)
    if not active:
        return ""

    header_row = rows[0]
    headers = []
    for ci in active:
        cell = header_row[ci] if ci < len(header_row) else ""
        cell = cell.replace("\n", " ").replace("|", r"\|").strip()
        headers.append(cell if cell else f"Col{ci + 1}")

    lines = ["| " + " | ".join(headers) + " |"]
    lines.append("| " + " | ".join(["---"] * len(headers)) + " |")

    for row in rows[1:]:
        cells = []
        for ci in active:
            cell = row[ci] if ci < len(row) else ""
            cell = cell.replace("|", r"\|").replace("\n", "<br>")
            cells.append(cell)
        lines.append("| " + " | ".join(cells) + " |")

    return "\n".join(lines)


def render_plain_section(rows, max_col):
    """
    Render as plain text. Non-empty cells joined with ' | ', one line per row.
    """
    lines = []
    for row in rows:
        parts = [row[ci] for ci in range(max_col) if ci < len(row) and row[ci]]
        if parts:
            lines.append(" | ".join(parts))
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Merged cell handling
# ---------------------------------------------------------------------------

def get_merged_cell_value(ws, row, col):
    """
    If the cell at (row, col) belongs to a merged range, return the value of
    the top-left cell in that range.
    """
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and
                merged_range.min_col <= col <= merged_range.max_col):
            top_left = ws.cell(merged_range.min_row, merged_range.min_col)
            return format_cell_value(top_left.value)
    return None


def is_merge_secondary(ws, row, col):
    """
    Return True if (row, col) is inside a merged range but is NOT the top-left cell.
    """
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and
                merged_range.min_col <= col <= merged_range.max_col):
            if row != merged_range.min_row or col != merged_range.min_col:
                return True
    return False


# ---------------------------------------------------------------------------
# Core logic
# ---------------------------------------------------------------------------

def collect_row_data(ws):
    """
    Iterate the worksheet, resolve merged cells, and return a list of rows.
    Each row is a list of string cell values (None for secondary merged cells).
    """
    all_rows = []
    for row in ws.iter_rows():
        row_data = []
        for cell in row:
            r, c = cell.row, cell.column
            if is_merge_secondary(ws, r, c):
                row_data.append(None)
            else:
                val = format_cell_value(cell.value)
                if not val:
                    # Try to get value from merged range top-left
                    merged_val = get_merged_cell_value(ws, r, c)
                    row_data.append(merged_val if merged_val else "")
                else:
                    row_data.append(val)
        all_rows.append(row_data)
    return all_rows


def group_into_sections(all_rows):
    """
    Split rows into sections by blank rows.
    A blank row has every cell None or empty string.
    Returns list of non-empty sections (each a list of rows).
    """
    sections = []
    current = []
    for row in all_rows:
        if all((c is None or c == "") for c in row):
            if current:
                sections.append(current)
                current = []
        else:
            # Strip trailing Nones
            stripped = list(row)
            while stripped and (stripped[-1] is None or stripped[-1] == ""):
                stripped.pop()
            # Replace None with ""
            stripped = [c if c is not None else "" for c in stripped]
            current.append(stripped)
    if current:
        sections.append(current)
    return sections


def process_sheet(ws):
    """
    Process a worksheet: classify each section and render as markdown.
    Returns a markdown string for the entire sheet.
    """
    all_rows = collect_row_data(ws)
    sections = group_into_sections(all_rows)

    parts = []
    for section in sections:
        max_col = max((len(r) for r in section), default=0)
        section_type = detect_section_type(section)

        if section_type == "key_value":
            rendered = render_key_value_section(section, max_col)
        elif section_type == "data_table":
            rendered = render_data_table(section, max_col)
        else:
            rendered = render_plain_section(section, max_col)

        if rendered.strip():
            parts.append(rendered)

    return "\n\n".join(parts)


def extract_images(xlsx_path, images_dir, prefix=""):
    """
    Treat xlsx as zip, iterate xl/media/*, copy each to images_dir.
    When prefix is non-empty, prepend it to each filename to avoid collisions
    across multiple source files.
    On BadZipFile: log warning and return 0.
    Returns count of extracted images.
    """
    os.makedirs(images_dir, exist_ok=True)
    count = 0
    try:
        with zipfile.ZipFile(xlsx_path, "r") as zf:
            for name in zf.namelist():
                if name.startswith("xl/media/"):
                    basename = os.path.basename(name)
                    if basename:
                        if prefix:
                            basename = f"{prefix}_{basename}"
                        target = os.path.join(images_dir, basename)
                        with zf.open(name) as src, open(target, "wb") as dst:
                            shutil.copyfileobj(src, dst)
                        count += 1
    except zipfile.BadZipFile:
        print(f"WARNING: Could not open {xlsx_path} as zip archive — skipping image extraction.")
    return count


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def parse_args():
    parser = argparse.ArgumentParser(
        description="Convert BD source .xlsx file(s) to BD.md with extracted images."
    )
    parser.add_argument("--input", required=True, action="append", dest="inputs",
                        help="Path to source .xlsx input file (can be specified multiple times for multi-file merge)")
    parser.add_argument("--output", required=True, help="Path to BD.md output file")
    parser.add_argument(
        "--images-dir", required=True, dest="images_dir",
        help="Directory to extract embedded images into"
    )
    return parser.parse_args()


def main():
    args = parse_args()

    # Load workbook with data_only=True for calculated values
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl is not installed. Run: pip install openpyxl")
        sys.exit(1)

    # Ensure output directory exists
    out_dir = os.path.dirname(os.path.abspath(args.output))
    os.makedirs(out_dir, exist_ok=True)

    md_parts = []
    total_images = 0
    total_sections = 0

    for input_path in args.inputs:
        if not os.path.exists(input_path):
            print(f"ERROR: Input file not found: {input_path}")
            sys.exit(1)

        # Extract images from this file (prefix with stem when multiple inputs)
        prefix = Path(input_path).stem if len(args.inputs) > 1 else ""
        image_count = extract_images(input_path, args.images_dir, prefix=prefix)
        total_images += image_count
        print(f"Images extracted from {input_path}: {image_count}")

        try:
            wb = openpyxl.load_workbook(input_path, data_only=True)
        except Exception as e:
            print(f"ERROR: Could not open workbook {input_path}: {e}")
            sys.exit(1)

        sheets = wb.sheetnames
        if not sheets:
            print(f"WARNING: Workbook {input_path} has no sheets — skipping.")
            continue

        print(f"Sheets in {input_path}: {len(sheets)}")

        for sheet_name in sheets:
            ws = wb[sheet_name]
            md_parts.append(f"# {sheet_name}")
            sheet_md = process_sheet(ws)
            if sheet_md.strip():
                section_count = sheet_md.count("\n\n") + 1
                total_sections += section_count
                md_parts.append(sheet_md)

    # Write markdown
    content = "\n\n".join(md_parts)

    # Append embedded images section if any
    if total_images > 0:
        image_bullets = []
        for fname in sorted(os.listdir(args.images_dir)):
            image_bullets.append(f"- ![{fname}](images/{fname})")
        images_section = "---\n## Embedded Images\n\n" + "\n".join(image_bullets)
        content = content + "\n\n" + images_section

    with open(args.output, "w", encoding="utf-8") as f:
        f.write(content)

    print(f"Output: {args.output}")
    print(f"Inputs: {len(args.inputs)} | Sections: ~{total_sections} | Images: {total_images}")


if __name__ == "__main__":
    main()
