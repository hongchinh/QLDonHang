#!/usr/bin/env python3
"""
export-tc-to-xlsx.py — Template-driven TC markdown → DJP xlsx export

Self-contained script (no external imports from scripts/).
Dependencies: openpyxl, pyyaml (see requirements.txt).

Usage:
    python3 export-tc-to-xlsx.py \
        --config <config.yaml> \
        --template <template.xlsx> \
        --gui <gui-testcases.md> \
        --function <function-testcases.md> \
        --output <output.xlsx>
"""

import argparse
import copy
import os
import re
import sys


# ---------------------------------------------------------------------------
# Global settings loader
# ---------------------------------------------------------------------------

_SETTINGS_PATH = os.path.join(os.path.dirname(__file__), "..", "settings.yaml")

def _load_settings():
    """Load skill-level settings.yaml. Prints warning and returns empty dict on failure."""
    try:
        import yaml
    except ImportError:
        print("WARNING: pyyaml not installed — settings.yaml not loaded, using empty defaults.", file=sys.stderr)
        return {}
    if not os.path.exists(_SETTINGS_PATH):
        print(f"WARNING: settings.yaml not found at {_SETTINGS_PATH} — using empty defaults.", file=sys.stderr)
        return {}
    with open(_SETTINGS_PATH, "r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}

_SETTINGS = _load_settings()


def _get(path, default=None):
    """Get a nested value from settings by dot-separated path."""
    node = _SETTINGS
    for key in path.split("."):
        if isinstance(node, dict):
            node = node.get(key)
        else:
            return default
    return node if node is not None else default


# ---------------------------------------------------------------------------
# Config-driven export defaults (loaded from settings.yaml)
# ---------------------------------------------------------------------------

GUI_DEFAULT_COLUMN_MAP = _get("export.column_defaults.gui", {})
FUNC_DEFAULT_COLUMN_MAP = _get("export.column_defaults.function", {})

# Backward compatibility alias
DEFAULT_COLUMN_MAP = dict(FUNC_DEFAULT_COLUMN_MAP)

# Header text → field key mapping for auto-detection from template row
HEADER_TEXT_MAP = _get("export.header_texts", {})

DEFAULT_SHEET_NAMES = _get("export.sheet_names", {})

HEADER_ROW = _get("export.header_row", 5)
MAX_COLUMN_FALLBACK = _get("export.max_column_fallback", 17)
MAX_COLUMN_SCAN = _get("export.max_column_scan", 20)
SECTION_DETECTION = _get("export.section_detection", {})
FORMULA_SCAN_MAX_ROW = _get("export.formula_scan_max_row", 4)
FALLBACK_ROWS = _get("export.fallback_rows", {"section": 7, "subsection": 8, "tc": 9})
TC_ID_PREFIXES = _get("tc_id.prefixes", {"gui": "GUI_", "function": "FUNC_", "common": "Common_"})
DEFAULT_GROUP_CODE = _get("tc_id.default_group_code", "N/A")

# Literal placeholders inside template content that are NOT meant to be replaced
# from config (e.g. "{M}" in Common pagination test descriptions). Suppress
# warnings for these to keep export logs clean.
LITERAL_PLACEHOLDERS = {"M"}


def detect_column_map_from_header(ws, header_row=None):
    """Auto-detect column map by scanning the header row for known header texts."""
    if header_row is None:
        header_row = HEADER_ROW
    detected = {}
    for col in range(1, (ws.max_column or MAX_COLUMN_SCAN) + 1):
        val = ws.cell(header_row, col).value
        if isinstance(val, str):
            val_stripped = val.strip()
            for header_text, field_key in HEADER_TEXT_MAP.items():
                if header_text.lower() in val_stripped.lower():
                    detected[field_key] = col
                    break
    return detected


def resolve_column_map(config, sheet_key="function", ws=None, header_row=None):
    """Merge config export_column_map over per-sheet defaults, with optional auto-detect."""
    # Start with per-sheet default
    if sheet_key == "gui":
        col_map = dict(GUI_DEFAULT_COLUMN_MAP)
    else:
        col_map = dict(FUNC_DEFAULT_COLUMN_MAP)

    # Apply per-sheet config override (new format: export_column_map.gui / export_column_map.function)
    custom = config.get("export_column_map")
    if custom and isinstance(custom, dict):
        per_sheet = custom.get(sheet_key)
        if per_sheet and isinstance(per_sheet, dict):
            for key, col_num in per_sheet.items():
                if isinstance(col_num, int) and key in col_map:
                    col_map[key] = col_num
        elif not per_sheet:
            # Backward compat: flat export_column_map (no per-sheet nesting)
            for key, col_num in custom.items():
                if isinstance(col_num, int) and key in col_map:
                    col_map[key] = col_num

    # Auto-detect from header row as reliability layer
    if ws is not None:
        detected = detect_column_map_from_header(ws, header_row)
        if detected:
            for key, col_num in detected.items():
                if key in col_map:
                    col_map[key] = col_num

    return col_map


def resolve_sheet_names(config):
    """Merge config export_sheets over defaults."""
    names = dict(DEFAULT_SHEET_NAMES)
    custom = config.get("export_sheets")
    if custom and isinstance(custom, dict):
        for key, name in custom.items():
            if isinstance(name, str) and key in names:
                names[key] = name
    return names


def resolve_environments(config):
    """
    Resolve test environment list using the chain:
      1. config['environments'] (if non-empty list)
      2. settings.environments[mode] (if mode is set and has defaults)
      3. settings.environments.all (ultimate fallback)
    """
    envs = config.get("environments", [])
    if envs and isinstance(envs, list):
        return envs
    mode = config.get("mode", "")
    if mode:
        mode_envs = _get(f"environments.{mode}", [])
        if mode_envs:
            return mode_envs
    return _get("environments.all", ["IOS/Chrome", "Android/Safari"])


def detect_env_blocks(ws):
    """
    Detect environment column blocks in the sub-header row (row below HEADER_ROW).
    Each block contains env status columns for one Test Round, delimited by
    "Execution Date" columns.

    Returns list of dicts with:
      - start_col: first env column in this block
      - env_names: list of environment name strings
      - post_env_col: "Execution Date" column after this block
    """
    sub_header_row = HEADER_ROW + 1
    shared_headers = {"Execution Date", "Assignee", "Bug ID"}
    data_headers = set(HEADER_TEXT_MAP.keys())

    blocks = []
    current_envs = []
    current_start = None

    for col in range(1, (ws.max_column or MAX_COLUMN_SCAN) + 1):
        val = ws.cell(sub_header_row, col).value
        stripped = val.strip() if isinstance(val, str) else ""

        if not stripped:
            continue

        if stripped == "Execution Date":
            if current_envs:
                blocks.append({
                    'start_col': current_start,
                    'env_names': current_envs,
                    'post_env_col': col,
                })
                current_envs = []
                current_start = None
        elif stripped in shared_headers:
            continue
        elif stripped in data_headers:
            continue
        else:
            if current_start is None:
                current_start = col
            current_envs.append(stripped)

    return blocks


def _extract_formula_range(ws, row, col):
    """Extract (start_row, end_row) from first range reference in a formula cell."""
    val = ws.cell(row, col).value
    if isinstance(val, str) and val.startswith("="):
        m = re.search(r'[A-Z]+(\d+):[A-Z]+(\d+)', val)
        if m:
            return int(m.group(1)), int(m.group(2))
    return None, None


def adjust_env_columns(ws, environments):
    """
    Adjust environment status columns in a TC sheet to match the environments list.
    Handles both Test Round 1 and Test Round 2 blocks.

    Returns dict with env_count, r1_env_cols, r2_env_cols, untested_cols,
    data_start_row, data_end_row. Returns None if no env columns detected.
    """
    blocks = detect_env_blocks(ws)
    if not blocks:
        return None

    n_new = len(environments)
    sub_header_row = HEADER_ROW + 1

    # Capture formula range before any modifications
    data_start_row, data_end_row = _extract_formula_range(ws, 2, 5)
    if data_start_row is None:
        data_start_row, data_end_row = 9, 2006

    # Unmerge any merges in summary rows (1-4) at stat area (col 5+) before
    # column operations — prevents MergedCell errors when columns shift
    summary_merges = [str(mr) for mr in ws.merged_cells.ranges
                      if mr.min_row <= 4 and mr.min_col >= 5]
    for mr_str in summary_merges:
        ws.unmerge_cells(mr_str)

    # Process blocks right-to-left to avoid offset issues
    for block in reversed(blocks):
        n_old = len(block['env_names'])
        delta = n_new - n_old
        first_col = block['start_col']

        if delta > 0:
            ws.insert_cols(first_col + n_old, delta)
            # Copy cell styles from first env col to newly inserted columns
            for row in range(1, sub_header_row + 6):
                src = ws.cell(row, first_col)
                for j in range(delta):
                    copy_cell_style(src, ws.cell(row, first_col + n_old + j))
        elif delta < 0:
            ws.delete_cols(first_col + n_new, abs(delta))

        # Update env headers in sub-header row
        for i, env_name in enumerate(environments):
            col = first_col + i
            ws.cell(sub_header_row, col).value = env_name
            if i > 0:
                copy_cell_style(ws.cell(sub_header_row, first_col),
                                ws.cell(sub_header_row, col))

    # Re-detect blocks to get new positions after adjustment
    new_blocks = detect_env_blocks(ws)

    r1_env_cols = []
    r2_env_cols = []
    untested_cols = []

    if len(new_blocks) >= 1:
        b = new_blocks[0]
        r1_env_cols = [b['start_col'] + i for i in range(n_new)]
        untested_cols.extend(r1_env_cols)
    if len(new_blocks) >= 2:
        b = new_blocks[1]
        r2_env_cols = [b['start_col'] + i for i in range(n_new)]
        untested_cols.extend(r2_env_cols)

    return {
        'env_count': n_new,
        'r1_env_cols': r1_env_cols,
        'r2_env_cols': r2_env_cols,
        'untested_cols': untested_cols,
        'data_start_row': data_start_row,
        'data_end_row': data_end_row,
        'environments': environments,
    }


def expand_summary_rows(ws, environments):
    """
    Insert extra summary rows in TC sheet when more than 2 environments.

    Original layout: row 2 = env1, row 3 = env2, row 4 = Total.
    For N > 2: insert (N - 2) rows after row 3, shifting Total and everything below.

    Returns row_delta (number of rows inserted, 0 when N <= 2).
    """
    n_env = len(environments)
    row_delta = max(0, n_env - 2)

    if row_delta == 0:
        return 0

    # Insert rows after row 3 (pushes row 4+ down)
    ws.insert_rows(4, row_delta)

    # Copy styles from row 3 to the newly inserted rows
    max_col = ws.max_column or MAX_COLUMN_FALLBACK
    for new_row_offset in range(row_delta):
        target_row = 4 + new_row_offset
        for col in range(1, max_col + 1):
            copy_cell_style(ws.cell(3, col), ws.cell(target_row, col))

    return row_delta


def rebuild_summary_formulas(ws, env_info, row_delta=0):
    """
    Rebuild COUNTIF/SUM formulas in summary rows after environment column adjustment.

    Writes one summary row per environment (rows 2 through 1+N),
    then a Total row at row 2+N.
    """
    from openpyxl.utils import get_column_letter

    r1_cols = env_info['r1_env_cols']
    r2_cols = env_info['r2_env_cols']
    n_env = env_info['env_count']
    ds = env_info['data_start_row'] + row_delta
    de = env_info['data_end_row'] + row_delta
    environments = env_info.get('environments', [])

    if not r1_cols:
        return

    sc = 5  # stat columns start at E (col 5)

    # Row 1: stat labels
    stat_labels = ["Untested", "OK", "   NG", "N/A", "Total", "Number of Bugs"]
    for i, label in enumerate(stat_labels):
        ws.cell(1, sc + i).value = label

    # Bug ID column (shared across envs)
    r1_bug_col = r1_cols[-1] + 3
    r2_bug_col = r2_cols[-1] + 3 if r2_cols else None

    def _env_formulas(r1_col, r2_col, row, env_name):
        r1 = get_column_letter(r1_col)
        # Col D: env label
        ws.cell(row, 4).value = env_name
        # Col E: Untested
        ws.cell(row, sc).value = f'=COUNTIF({r1}{ds}:{r1}{de}, "Untested")'
        # Col F: OK
        ok = f'=COUNTIF({r1}{ds}:{r1}{de}, "OK")'
        ng = f'=COUNTIF({r1}{ds}:{r1}{de}, "NG")'
        if r2_col:
            r2 = get_column_letter(r2_col)
            ok += f'+COUNTIF({r2}{ds}:{r2}{de}, "OK")'
            ng += f'-COUNTIF({r2}{ds}:{r2}{de}, "OK")'
        ws.cell(row, sc + 1).value = ok
        # Col G: NG
        ws.cell(row, sc + 2).value = ng
        # Col H: N/A
        ws.cell(row, sc + 3).value = f'=COUNTIF({r1}{ds}:{r1}{de}, "N/A")'
        # Col I: Total
        ws.cell(row, sc + 4).value = f'=SUM(E{row}:H{row})'
        # Col J: Number of Bugs
        parts = [f'COUNTA({get_column_letter(r1_bug_col)}{ds}:{get_column_letter(r1_bug_col)}{de})']
        if r2_bug_col:
            parts.append(f'COUNTA({get_column_letter(r2_bug_col)}{ds}:{get_column_letter(r2_bug_col)}{de})')
        ws.cell(row, sc + 5).value = "=" + "+".join(parts)

    # Write per-env summary rows (rows 2 through 1+N)
    for i in range(n_env):
        row = 2 + i
        r1_col = r1_cols[i]
        r2_col = r2_cols[i] if i < len(r2_cols) else None
        env_name = environments[i] if i < len(environments) else f"Env {i+1}"
        _env_formulas(r1_col, r2_col, row, env_name)

    # Total row at row 2 + N
    total_row = 2 + n_env
    ws.cell(total_row, 4).value = "Total"
    for i in range(5):
        col_letter = get_column_letter(sc + i)
        ws.cell(total_row, sc + i).value = f'=SUM({col_letter}2:{col_letter}{total_row - 1})'

    # Clear stale values below the Total row (handles N < previous N).
    # Bound to rows strictly above HEADER_ROW so the data header row (and merged
    # header cells like 'Test Steps' / 'Expected Result') is never wiped.
    from openpyxl.cell.cell import MergedCell
    for clear_row in range(total_row + 1, min(total_row + 3, HEADER_ROW)):
        for c in range(4, sc + 6):
            cell = ws.cell(clear_row, c)
            if not isinstance(cell, MergedCell):
                cell.value = None


def adjust_test_report(ws, environments, sheet_names):
    """
    Adjust Test Report sheet to match dynamic environments.

    Each environment gets 4 columns: Untested, OK, NG, N/A.
    After all env groups: Total column with SUM.
    Formulas reference TC sheet summary rows (row 2+i for i-th environment).
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font

    n_env = len(environments)
    data_start_col = 4  # Column D — first data column in Test Report
    old_data_cols = 8  # Original: 2 envs × 4 stats
    new_data_cols = n_env * 4
    delta = new_data_cols - old_data_cols

    total_col = data_start_col + new_data_cols

    # Remove old merges in rows 9-10 BEFORE column adjustment
    merges_to_remove = [str(mr) for mr in ws.merged_cells.ranges
                        if mr.min_row >= 9 and mr.min_row <= 10
                        and mr.min_col >= data_start_col]
    for mr_str in merges_to_remove:
        ws.unmerge_cells(mr_str)

    # Insert or delete columns
    if delta > 0:
        ws.insert_cols(data_start_col + old_data_cols, delta)
    elif delta < 0:
        ws.delete_cols(data_start_col + new_data_cols, abs(delta))

    # Clear rows 9-10 data area
    for col in range(data_start_col, total_col + 2):
        ws.cell(9, col).value = None
        ws.cell(10, col).value = None

    # Row 9: env group headers (merged 4 cols each)
    for i, env_name in enumerate(environments):
        start_col = data_start_col + i * 4
        end_col = start_col + 3
        ws.merge_cells(start_row=9, start_column=start_col,
                       end_row=9, end_column=end_col)
        cell = ws.cell(9, start_col)
        cell.value = env_name
        cell.font = Font(bold=True)

    # Total header (merge rows 9-10)
    ws.merge_cells(start_row=9, start_column=total_col,
                   end_row=10, end_column=total_col)
    ws.cell(9, total_col).value = "Total number of test cases"

    # Row 10: stat sub-headers
    stat_labels = ["Untested", "OK", "NG", "N/A"]
    for i in range(n_env):
        for j, label in enumerate(stat_labels):
            ws.cell(10, data_start_col + i * 4 + j).value = label

    # Rows 11-13: formula references to TC sheet summary stats
    tc_sheet_refs = [
        (11, "function", "FUNC"),
        (12, "gui", "UI"),
        (13, "common", "COMMON"),
    ]

    stat_col_start = 5  # Column E in TC sheets (summary stats always at E-H)

    for report_row, sheet_key, label in tc_sheet_refs:
        ws.cell(report_row, 3).value = label
        tc_sheet_name = sheet_names.get(sheet_key, "")

        for i in range(n_env):
            summary_row = 2 + i

            for j in range(4):  # Untested, OK, NG, N/A
                report_col = data_start_col + i * 4 + j
                tc_stat_letter = get_column_letter(stat_col_start + j)
                ws.cell(report_row, report_col).value = (
                    f"={tc_sheet_name}!{tc_stat_letter}{summary_row}"
                )

        # Total formula
        total_start = get_column_letter(data_start_col)
        total_end = get_column_letter(total_col - 1)
        ws.cell(report_row, total_col).value = (
            f"=SUM({total_start}{report_row}:{total_end}{report_row})"
        )

    # Row 14: Total row
    ws.cell(14, 3).value = "Total"
    for col in range(data_start_col, total_col + 1):
        col_letter = get_column_letter(col)
        ws.cell(14, col).value = f"=SUM({col_letter}11:{col_letter}13)"

    # Coverage formulas (rows 16-17)
    ok_refs = []
    ng_refs = []
    na_refs = []
    for i in range(n_env):
        ok_refs.append(f"{get_column_letter(data_start_col + i * 4 + 1)}14")
        ng_refs.append(f"{get_column_letter(data_start_col + i * 4 + 2)}14")
        na_refs.append(f"{get_column_letter(data_start_col + i * 4 + 3)}14")

    ok_sum = "+".join(ok_refs)
    ng_sum = "+".join(ng_refs)
    na_sum = "+".join(na_refs)
    total_ref = f"{get_column_letter(total_col)}14"

    # Test coverage = (OK + NG) / (Total - N/A) * 100
    ws.cell(16, 5).value = f"=({ok_sum}+{ng_sum})*100/({total_ref}-({na_sum}))"
    # Test success coverage = OK / (Total - N/A) * 100
    ws.cell(17, 5).value = f"=({ok_sum})*100/({total_ref}-({na_sum}))"


# ---------------------------------------------------------------------------
# Markdown TC parser (inlined from old parse-markdown-testcases.py)
# ---------------------------------------------------------------------------

RE_GROUP = re.compile(r"^## Group:\s+(\S+)(?:\s*—\s*(.+?))?\s*$")
RE_SECTION = re.compile(r"^## ([A-Z]\..*)")
RE_SUBSECTION = re.compile(r"^### \*{0,2}(?:GUI |FUNC )?([A-Z](?:\.\d+)+)\s+(.+?)\*{0,2}\s*$")
RE_TC_ID = re.compile(r"^### ((?:GUI_|FUNC_)?[A-Z](?:\.\d+)+_\d+)\s*$")
RE_PRECONDITION = re.compile(r"^> Pre-condition:\s*(.*)")
RE_FIELD = re.compile(r"^- \*\*(\w+):\*\*\s*(.*)")
RE_TEMPLATE_SECTION_ID = re.compile(r"^(?:GUI |FUNC |Common )?[A-Z](?:\.\d+)+$")
RE_SHEET_TC_ID = re.compile(r"^(?:(?:GUI|FUNC|Common)_)?(?:[A-Z](?:\.\d+)+_\d+|\d+_\d+)$")

FIELD_MAP = {
    "precondition": "precondition",
    "title": "title",
    "steps": "steps",
    "expected": "expected",
}

def parse_markdown_tcs(md_path):
    """
    Parse a markdown TC file and return a list of entry tuples.

    Entry tuple formats:
      ("group", code, label)
      ("section", text, "")
      ("subsection", id, description)
      ("precondition", text, "")
      ("tc", id, {fields_dict})

    Parser ordering:
      - Group before Section
      - Subsection only when TC ID regex does not also match
    """
    if not os.path.exists(md_path):
        return []

    entries = []
    current_tc = None
    current_field = None
    group_emitted = False

    def flush_tc():
        nonlocal current_tc, current_field
        if current_tc is not None:
            tc_id, fields = current_tc
            fields = {k: v.strip() if isinstance(v, str) else v for k, v in fields.items()}
            entries.append(("tc", tc_id, fields))
            current_tc = None
            current_field = None

    def maybe_inject_implicit_group():
        nonlocal group_emitted
        if not group_emitted:
            entries.append(("group", DEFAULT_GROUP_CODE, ""))
            group_emitted = True

    with open(md_path, "r", encoding="utf-8") as f:
        lines = f.readlines()

    for line in lines:
        line = line.rstrip("\n")

        # Skip top-level headings
        if line.startswith("# ") and not line.startswith("## "):
            flush_tc()
            continue

        # Skip HTML comments
        if line.startswith("<!--") or line.startswith("-->"):
            continue

        # Group (check before Section)
        m = RE_GROUP.match(line)
        if m:
            flush_tc()
            code = m.group(1)
            label = m.group(2) or ""
            entries.append(("group", code, label.strip()))
            group_emitted = True
            continue

        # Section
        m = RE_SECTION.match(line)
        if m:
            flush_tc()
            maybe_inject_implicit_group()
            entries.append(("section", m.group(1).strip(), ""))
            continue

        # TC ID (check before Subsection)
        m = RE_TC_ID.match(line)
        if m:
            flush_tc()
            maybe_inject_implicit_group()
            tc_id = m.group(1)
            current_tc = (tc_id, {})
            current_field = None
            continue

        # Subsection (only when TC ID regex does not match — already checked above)
        m = RE_SUBSECTION.match(line)
        if m:
            flush_tc()
            maybe_inject_implicit_group()
            entries.append(("subsection", m.group(1), m.group(2).strip()))
            continue

        # Pre-condition blockquote
        m = RE_PRECONDITION.match(line)
        if m:
            flush_tc()
            maybe_inject_implicit_group()
            entries.append(("precondition", m.group(1).strip(), ""))
            continue

        # Field line (inside a TC)
        if current_tc is not None:
            m = RE_FIELD.match(line)
            if m:
                raw_key = m.group(1).lower()
                value = m.group(2)
                mapped = FIELD_MAP.get(raw_key, raw_key)
                current_tc[1][mapped] = value
                current_field = mapped
                continue

            # Continuation line (inside a TC)
            if current_field is not None:
                if line.strip():
                    current_tc[1][current_field] = current_tc[1].get(current_field, "") + "\n" + line.strip()
                    continue
                else:
                    # Blank line inside TC — preserve paragraph breaks within Steps/Expected.
                    # Trailing blank lines are trimmed in flush_tc().
                    current_tc[1][current_field] = current_tc[1].get(current_field, "") + "\n"
                    continue

    flush_tc()

    # Post-parse: remove implicit N/A groups when real groups exist
    real_groups = [e for e in entries if e[0] == "group" and e[1] != DEFAULT_GROUP_CODE]
    if real_groups:
        entries = [e for e in entries if not (e[0] == "group" and e[1] == DEFAULT_GROUP_CODE)]

    return entries


# ---------------------------------------------------------------------------
# Template operations
# ---------------------------------------------------------------------------


def replace_placeholders(wb, config, doc_code):
    """
    Scan every cell in every sheet.
    Replace {key} and {{key}} with config value.
    Do NOT alter formula cells (cells starting with '=').
    Log warnings for unknown placeholders.
    """
    all_keys = dict(config)
    all_keys["doc_code"] = doc_code

    placeholder_re = re.compile(r"\{\{?(\w+)\}?\}")

    def replace_in_value(value):
        if not isinstance(value, str):
            return value
        if value.startswith("="):
            return value  # Do not alter formulas

        def replacer(m):
            key = m.group(1)
            if key in all_keys:
                return str(all_keys[key])
            if key in LITERAL_PLACEHOLDERS:
                return m.group(0)  # Leave unchanged silently
            print(f"WARNING: Unknown placeholder: {{{key}}}")
            return m.group(0)  # Leave unchanged

        return placeholder_re.sub(replacer, value)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    new_val = replace_in_value(cell.value)
                    if new_val != cell.value:
                        cell.value = new_val



def find_data_start_rows(ws):
    """
    Identify the section header row, subsection row, and TC data row.

    Strategy (in order):
    1. Look for a row with a merged cell spanning ≥ 10 columns (DJP template pattern)
    2. Look for a row where col A contains text matching a section pattern (e.g. "A. ...")
    3. Fallback to (7, 8, 9)

    After finding the section header, scan nearby rows by content pattern to detect
    the subsection row and TC data row (handles templates with extra rows like
    Pre-condition between section header and subsection).
    """
    fb = FALLBACK_ROWS
    fb_section, fb_sub, fb_tc = fb.get("section", 7), fb.get("subsection", 8), fb.get("tc", 9)
    sd = SECTION_DETECTION
    merge_span_min = sd.get("merge_span_min", 10)
    merged_scan_start = sd.get("merged_scan_start", 6)
    merged_scan_end = sd.get("merged_scan_end", 15)
    pattern_scan_start = sd.get("pattern_scan_start", 5)
    pattern_scan_end = sd.get("pattern_scan_end", 20)
    subsection_window = sd.get("subsection_window", 6)

    max_row = ws.max_row
    if max_row < fb_tc:
        return fb_section, fb_sub, fb_tc

    section_row_idx = None

    for r in range(merged_scan_start, min(max_row, merged_scan_end)):
        for merged in ws.merged_cells.ranges:
            if merged.min_row == r and merged.max_row == r:
                span = merged.max_col - merged.min_col + 1
                if span >= merge_span_min:
                    section_row_idx = r
                    break
        if section_row_idx is not None:
            break

    if section_row_idx is None:
        section_re = re.compile(r"^[A-Z]\.\s")
        for r in range(pattern_scan_start, min(max_row, pattern_scan_end)):
            val = ws.cell(r, 1).value
            if isinstance(val, str) and section_re.match(val.strip()):
                section_row_idx = r
                break

    if section_row_idx is None:
        return fb_section, fb_sub, fb_tc

    subsection_re = re.compile(r"^(?:FUNC |GUI |Common )?[A-Z]\.\d+")
    tc_re = re.compile(r"^(?:FUNC_|GUI_|Common_)[\w.]+_\d+")

    subsection_row = None
    tc_row = None

    for r in range(section_row_idx + 1, min(section_row_idx + subsection_window, max_row + 1)):
        val = ws.cell(r, 1).value
        if not isinstance(val, str):
            continue
        val_stripped = val.strip()
        if subsection_row is None and subsection_re.match(val_stripped):
            subsection_row = r
        if tc_row is None and tc_re.match(val_stripped):
            tc_row = r

    if subsection_row is None or tc_row is None:
        return section_row_idx, section_row_idx + 1, section_row_idx + 2

    return section_row_idx, subsection_row, tc_row



def copy_cell_style(src_cell, dst_cell):
    """Copy font, fill, border, alignment, number_format from src to dst."""
    if src_cell.has_style:
        dst_cell.font = copy.copy(src_cell.font)
        dst_cell.fill = copy.copy(src_cell.fill)
        dst_cell.border = copy.copy(src_cell.border)
        dst_cell.alignment = copy.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format



def expand_formula_ranges(ws, last_data_row, summary_end_row=None):
    """
    After filling, check COUNTIF/COUNTA formulas in summary rows.
    If the last data row exceeds the formula range end, expand to cover all inserted rows.
    """
    range_end_re = re.compile(r"([A-Z]+)(\d+):([A-Z]+)(\d+)")
    scan_end = summary_end_row if summary_end_row else FORMULA_SCAN_MAX_ROW

    for r in range(1, scan_end + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(r, col)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula = cell.value

                def expand_match(m):
                    col1, row1, col2, row2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
                    # Only expand vertical data ranges — skip horizontal ranges (E2:H2)
                    # and summary-area ranges (I2:I3) that start above the header row
                    if col1 == col2 and row2 < last_data_row and row1 >= HEADER_ROW:
                        return f"{col1}{row1}:{col2}{last_data_row}"
                    return m.group(0)

                new_formula = range_end_re.sub(expand_match, formula)
                if new_formula != formula:
                    cell.value = new_formula



def resolve_preserve_sections(config, sheet_name):
    """Return normalized section IDs to preserve from the template for a sheet."""
    raw = config.get("template_preserve_sections")

    if raw is None:
        return []

    if not isinstance(raw, dict):
        raise ValueError("template_preserve_sections must be a mapping of sheet name to section ID list")

    sections = raw.get(sheet_name, [])
    if sections is None:
        return []

    if not isinstance(sections, list) or any(not isinstance(item, str) for item in sections):
        raise ValueError(f"template_preserve_sections.{sheet_name} must be a list of section IDs")

    normalized = []
    for item in sections:
        section_id = item.strip()
        if section_id and section_id not in normalized:
            normalized.append(section_id)
    return normalized



def capture_row_snapshot(ws, row_idx, max_col):
    """Capture cell values/styles for one row."""
    row = []
    for col in range(1, max_col + 1):
        src = ws.cell(row_idx, col)
        row.append({
            "value": src.value,
            "style": copy.copy(src._style),
        })
    return row



def capture_template_section_blocks(ws, section_row_idx, preserve_section_ids, max_col):
    """Capture configured template subsection blocks (e.g. A.1, A.2) before clearing the sheet."""
    if not preserve_section_ids:
        return []

    section_rows = []
    for row_idx in range(section_row_idx, ws.max_row + 1):
        value = ws.cell(row_idx, 1).value
        if isinstance(value, str):
            raw = value.strip()
            if RE_TEMPLATE_SECTION_ID.match(raw):
                section_id = re.sub(r"^(?:GUI |FUNC |Common )", "", raw)
                section_rows.append((section_id, row_idx))

    section_map = {section_id: row_idx for section_id, row_idx in section_rows}
    missing = [section_id for section_id in preserve_section_ids if section_id not in section_map]
    if missing:
        raise ValueError(
            f"Sheet '{ws.title}' is missing template sections required by template_preserve_sections: {', '.join(missing)}"
        )

    blocks = []
    for section_id in preserve_section_ids:
        start_row = section_map[section_id]
        next_rows = [row_idx for _, row_idx in section_rows if row_idx > start_row]
        end_row = next_rows[0] - 1 if next_rows else ws.max_row

        merges = []
        for merged in ws.merged_cells.ranges:
            if start_row <= merged.min_row and merged.max_row <= end_row:
                merges.append({
                    "min_row_offset": merged.min_row - start_row,
                    "max_row_offset": merged.max_row - start_row,
                    "min_col": merged.min_col,
                    "max_col": merged.max_col,
                })

        blocks.append({
            "section_id": section_id,
            "rows": [capture_row_snapshot(ws, row_idx, max_col) for row_idx in range(start_row, end_row + 1)],
            "row_heights": [ws.row_dimensions[row_idx].height for row_idx in range(start_row, end_row + 1)],
            "merges": merges,
        })

    return blocks



def insert_template_block(ws, start_row, block):
    """Insert one captured template block at the target row."""
    row_count = len(block["rows"])
    if row_count == 0:
        return start_row

    ws.insert_rows(start_row, row_count)

    for row_offset, row in enumerate(block["rows"]):
        row_num = start_row + row_offset
        row_height = block["row_heights"][row_offset]
        if row_height is not None:
            ws.row_dimensions[row_num].height = row_height

        for col, captured in enumerate(row, start=1):
            dst = ws.cell(row_num, col)
            dst.value = captured["value"]
            dst._style = copy.copy(captured["style"])

    for merge in block["merges"]:
        ws.merge_cells(
            start_row=start_row + merge["min_row_offset"],
            end_row=start_row + merge["max_row_offset"],
            start_column=merge["min_col"],
            end_column=merge["max_col"],
        )

    return start_row + row_count



def insert_template_blocks(ws, start_row, blocks):
    """Insert multiple captured template blocks sequentially."""
    current_row = start_row
    for block in blocks:
        current_row = insert_template_block(ws, current_row, block)
    return current_row



def filter_generated_entries(entries, preserve_section_ids):
    """Remove generated subsections that are supplied directly by the template."""
    if not preserve_section_ids:
        return entries, []

    filtered = []
    skipped_sections = []
    skip_current_subsection = False

    for entry in entries:
        entry_type = entry[0]

        if entry_type == "subsection":
            subsection_id = entry[1]
            skip_current_subsection = subsection_id in preserve_section_ids
            if skip_current_subsection:
                if subsection_id not in skipped_sections:
                    skipped_sections.append(subsection_id)
                continue
            filtered.append(entry)
            continue

        if entry_type in ("group", "section"):
            skip_current_subsection = False
            filtered.append(entry)
            continue

        if skip_current_subsection:
            continue

        filtered.append(entry)

    return filtered, skipped_sections



def count_sheet_tcs(ws):
    """Count testcase rows currently present in a worksheet."""
    count = 0
    for row_idx in range(1, ws.max_row + 1):
        value = ws.cell(row_idx, 1).value
        if isinstance(value, str) and RE_SHEET_TC_ID.match(value.strip()):
            count += 1
    return count


# ---------------------------------------------------------------------------
# Sheet filling
# ---------------------------------------------------------------------------


def fill_tc_sheet(ws, entries, max_col, preserve_section_ids=None, col_map=None, tc_id_prefix="", subsection_prefix=True, untested_cols_override=None):
    """
    Fill a GUI or Function sheet with parsed TC entries.

    Strategy:
    1. Identify template anchor rows.
    2. Capture template subsection blocks configured for preservation.
    3. Delete all rows from the section area onwards.
    4. Rebuild the sheet with preserved blocks + generated entries.

    tc_id_prefix: If set (e.g. "GUI_", "FUNC_"), prepend to TC IDs that
    don't already carry the prefix.
    subsection_prefix: If True, prepend prefix label to subsection rows
    (e.g. "FUNC A.1"). If False, subsection rows show bare ID (e.g. "A.1").
    """
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    preserve_section_ids = preserve_section_ids or []
    if col_map is None:
        col_map = dict(DEFAULT_COLUMN_MAP)
    section_row_idx, subsection_row_idx, tc_row_idx = find_data_start_rows(ws)

    section_styles = {}
    subsection_styles = {}
    precondition_styles = {}
    tc_styles = {}
    for col in range(1, max_col + 1):
        sec_c = ws.cell(section_row_idx, col)
        if sec_c.has_style:
            section_styles[col] = {
                "font": copy.copy(sec_c.font),
                "fill": copy.copy(sec_c.fill),
                "border": copy.copy(sec_c.border),
                "alignment": copy.copy(sec_c.alignment),
                "number_format": sec_c.number_format,
            }
        sc = ws.cell(subsection_row_idx, col)
        if sc.has_style:
            subsection_styles[col] = {
                "font": copy.copy(sc.font),
                "fill": copy.copy(sc.fill),
                "border": copy.copy(sc.border),
                "alignment": copy.copy(sc.alignment),
                "number_format": sc.number_format,
            }
        tc_c = ws.cell(tc_row_idx, col)
        if tc_c.has_style:
            tc_styles[col] = {
                "font": copy.copy(tc_c.font),
                "fill": copy.copy(tc_c.fill),
                "border": copy.copy(tc_c.border),
                "alignment": copy.copy(tc_c.alignment),
                "number_format": tc_c.number_format,
            }
    # Capture precondition row style if an intermediate row exists between section and subsection
    if subsection_row_idx > section_row_idx + 1:
        pre_row = section_row_idx + 1
        for col in range(1, max_col + 1):
            pc = ws.cell(pre_row, col)
            if pc.has_style:
                precondition_styles[col] = {
                    "font": copy.copy(pc.font),
                    "fill": copy.copy(pc.fill),
                    "border": copy.copy(pc.border),
                    "alignment": copy.copy(pc.alignment),
                    "number_format": pc.number_format,
                }

    # Detect "Untested" columns from template TC row (or use override)
    if untested_cols_override is not None:
        untested_cols = list(untested_cols_override)
    else:
        untested_cols = []
        for col in range(1, max_col + 1):
            val = ws.cell(tc_row_idx, col).value
            if isinstance(val, str) and val.strip() == "Untested":
                untested_cols.append(col)

    template_blocks = capture_template_section_blocks(ws, section_row_idx, preserve_section_ids, max_col)

    rows_to_delete = ws.max_row - section_row_idx + 1
    if rows_to_delete > 0:
        ws.delete_rows(section_row_idx, rows_to_delete)

    merges_to_remove = [mr for mr in ws.merged_cells.ranges if mr.min_row >= section_row_idx]
    for mr in merges_to_remove:
        ws.merged_cells.remove(mr)

    filtered_entries, skipped_sections = filter_generated_entries(entries, preserve_section_ids)
    if skipped_sections:
        print(
            f"INFO: Sheet '{ws.title}' skipped generated subsections already preserved from template: {', '.join(skipped_sections)}"
        )

    groups = [entry for entry in filtered_entries if entry[0] == "group"]
    flat_layout = not groups or (len(groups) == 1 and groups[0][1] == DEFAULT_GROUP_CODE)

    def apply_styles(row_num, style_map, fallback_map=None):
        for col in range(1, max_col + 1):
            cell = ws.cell(row_num, col)
            styles = style_map.get(col) or (fallback_map.get(col) if fallback_map else None)
            if styles:
                cell.font = copy.copy(styles["font"])
                cell.fill = copy.copy(styles["fill"])
                cell.border = copy.copy(styles["border"])
                cell.alignment = copy.copy(styles["alignment"])
                cell.number_format = styles["number_format"]

    current_row = section_row_idx

    if flat_layout and template_blocks:
        current_row = insert_template_blocks(ws, current_row, template_blocks)

    for entry in filtered_entries:
        entry_type = entry[0]

        if entry_type == "group":
            if flat_layout:
                continue
            code, label = entry[1], entry[2]
            ws.insert_rows(current_row)
            apply_styles(current_row, section_styles, subsection_styles)
            group_text = f"{code} {label}".strip()
            ws.cell(current_row, 2).value = group_text
            current_row += 1

            if template_blocks:
                current_row = insert_template_blocks(ws, current_row, template_blocks)

        elif entry_type == "section":
            text = entry[1]
            ws.insert_rows(current_row)
            apply_styles(current_row, section_styles, subsection_styles)
            merge_ref = f"A{current_row}:{get_column_letter(max_col)}{current_row}"
            ws.merge_cells(merge_ref)
            c = ws.cell(current_row, 1)
            c.value = text
            c.font = Font(bold=True)
            current_row += 1

        elif entry_type == "subsection":
            sub_id, description = entry[1], entry[2]
            ws.insert_rows(current_row)
            apply_styles(current_row, subsection_styles)
            # Add prefix for sheet-specific subsections (e.g. "FUNC A.1")
            if tc_id_prefix and subsection_prefix:
                prefix_label = tc_id_prefix.rstrip("_")
                ws.cell(current_row, 1).value = f"{prefix_label} {sub_id}"
            else:
                ws.cell(current_row, 1).value = sub_id
            ws.cell(current_row, 2).value = description
            current_row += 1

        elif entry_type == "precondition":
            text = entry[1]
            ws.insert_rows(current_row)
            apply_styles(current_row, precondition_styles if precondition_styles else subsection_styles)
            merge_ref = f"A{current_row}:{get_column_letter(max_col)}{current_row}"
            ws.merge_cells(merge_ref)
            c = ws.cell(current_row, 1)
            c.value = f"Pre-condition: {text}"
            c.font = Font(italic=True)
            current_row += 1

        elif entry_type == "tc":
            tc_id, fields = entry[1], entry[2]
            if tc_id_prefix and not tc_id.startswith(tc_id_prefix):
                tc_id = tc_id_prefix + tc_id
            ws.insert_rows(current_row)
            apply_styles(current_row, tc_styles, subsection_styles)
            ws.cell(current_row, col_map["tc_id"]).value = tc_id
            # Only write fields that exist in this sheet's column map
            for field_key in ["title", "precondition", "steps", "expected"]:
                if field_key in col_map and fields.get(field_key):
                    ws.cell(current_row, col_map[field_key]).value = fields[field_key]
            for uc in untested_cols:
                ws.cell(current_row, uc).value = "Untested"
            current_row += 1

    return current_row - 1


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def parse_args():
    parser = argparse.ArgumentParser(
        description="Export TC markdown files to DJP xlsx using a template."
    )
    parser.add_argument("--config", required=True, help="Path to config.yaml")
    parser.add_argument("--template", required=True, help="Path to DJP_TestCase_Template_Ver1.0.xlsx")
    parser.add_argument("--gui", required=True, help="Path to gui-testcases.md")
    parser.add_argument("--function", required=True, help="Path to function-testcases.md")
    parser.add_argument("--common", default=None, help="Path to common-testcases.md (optional)")
    parser.add_argument("--output", required=True, help="Path for output .xlsx file")
    return parser.parse_args()



def load_config(config_path):
    try:
        import yaml
    except ImportError:
        print("ERROR: pyyaml is not installed. Run: pip install pyyaml")
        sys.exit(1)

    if not os.path.exists(config_path):
        print(f"ERROR: Config file not found: {config_path}")
        sys.exit(1)

    with open(config_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}



def main():
    args = parse_args()

    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl is not installed. Run: pip install openpyxl")
        sys.exit(1)

    try:
        import yaml  # noqa: F401
    except ImportError:
        print("ERROR: pyyaml is not installed. Run: pip install pyyaml")
        sys.exit(1)

    if not os.path.exists(args.template):
        print(f"ERROR: Template file not found: {args.template}")
        sys.exit(1)

    config = load_config(args.config)

    screen_id = config.get("screen_id", "")
    screen_name = str(config.get("screen_name", "")).replace(" ", "")
    version = config.get("version", "1.0")
    doc_code = f"{screen_id}_{screen_name}_Testcase_v{version}"

    try:
        wb = openpyxl.load_workbook(args.template)
    except Exception as e:
        print(f"ERROR: Could not open template: {e}")
        sys.exit(1)

    sheet_names = resolve_sheet_names(config)

    required_sheets = list(sheet_names.values())
    missing = [s for s in required_sheets if s not in wb.sheetnames]
    if missing:
        print(f"ERROR: Template missing required sheets: {', '.join(missing)}")
        sys.exit(1)

    replace_placeholders(wb, config, doc_code)

    environments = resolve_environments(config)

    gui_entries = parse_markdown_tcs(args.gui)
    func_entries = parse_markdown_tcs(args.function)
    common_entries = parse_markdown_tcs(args.common) if args.common else []

    if not gui_entries:
        print(f"WARNING: No entries parsed from GUI markdown: {args.gui}")
    if not func_entries:
        print(f"WARNING: No entries parsed from Function markdown: {args.function}")

    try:
        gui_preserve_sections = resolve_preserve_sections(config, sheet_names["gui"])
        func_preserve_sections = resolve_preserve_sections(config, sheet_names["function"])
        common_preserve_sections = resolve_preserve_sections(config, sheet_names["common"])
    except ValueError as e:
        print(f"ERROR: {e}")
        sys.exit(1)

    n_env = len(environments)
    summary_end_row = 2 + n_env

    try:
        ws_gui = wb[sheet_names["gui"]]
        env_info_gui = adjust_env_columns(ws_gui, environments)
        row_delta_gui = expand_summary_rows(ws_gui, environments)
        if env_info_gui:
            rebuild_summary_formulas(ws_gui, env_info_gui, row_delta_gui)
        effective_header_row = HEADER_ROW + row_delta_gui
        max_col_gui = ws_gui.max_column or MAX_COLUMN_FALLBACK
        col_map_gui = resolve_column_map(config, "gui", ws_gui, header_row=effective_header_row)
        if gui_entries or gui_preserve_sections:
            last_row_gui = fill_tc_sheet(ws_gui, gui_entries, max_col_gui, gui_preserve_sections, col_map_gui, tc_id_prefix=TC_ID_PREFIXES.get("gui", "GUI_"), subsection_prefix=True, untested_cols_override=env_info_gui['untested_cols'] if env_info_gui else None)
            expand_formula_ranges(ws_gui, last_row_gui, summary_end_row)

        ws_func = wb[sheet_names["function"]]
        env_info_func = adjust_env_columns(ws_func, environments)
        row_delta_func = expand_summary_rows(ws_func, environments)
        if env_info_func:
            rebuild_summary_formulas(ws_func, env_info_func, row_delta_func)
        effective_header_row_func = HEADER_ROW + row_delta_func
        max_col_func = ws_func.max_column or MAX_COLUMN_FALLBACK
        col_map_func = resolve_column_map(config, "function", ws_func, header_row=effective_header_row_func)
        if func_entries or func_preserve_sections:
            last_row_func = fill_tc_sheet(ws_func, func_entries, max_col_func, func_preserve_sections, col_map_func, tc_id_prefix=TC_ID_PREFIXES.get("function", "FUNC_"), untested_cols_override=env_info_func['untested_cols'] if env_info_func else None)
            expand_formula_ranges(ws_func, last_row_func, summary_end_row)

        ws_common = wb[sheet_names["common"]]
        env_info_common = adjust_env_columns(ws_common, environments)
        row_delta_common = expand_summary_rows(ws_common, environments)
        if env_info_common:
            rebuild_summary_formulas(ws_common, env_info_common, row_delta_common)
        if common_entries or common_preserve_sections:
            effective_header_row_common = HEADER_ROW + row_delta_common
            max_col_common = ws_common.max_column or MAX_COLUMN_FALLBACK
            col_map_common = resolve_column_map(config, "common", ws_common, header_row=effective_header_row_common)
            last_row_common = fill_tc_sheet(ws_common, common_entries, max_col_common, common_preserve_sections, col_map_common, tc_id_prefix=TC_ID_PREFIXES.get("common", "Common_"), untested_cols_override=env_info_common['untested_cols'] if env_info_common else None)
            expand_formula_ranges(ws_common, last_row_common, summary_end_row)
    except ValueError as e:
        print(f"ERROR: {e}")
        sys.exit(1)

    # Adjust Test Report to match dynamic environments
    ws_report = wb[sheet_names["test_report"]]
    adjust_test_report(ws_report, environments, sheet_names)

    gui_tc_count = count_sheet_tcs(wb[sheet_names["gui"]])
    func_tc_count = count_sheet_tcs(wb[sheet_names["function"]])
    common_tc_count = count_sheet_tcs(wb[sheet_names["common"]])

    out_dir = os.path.dirname(os.path.abspath(args.output))
    os.makedirs(out_dir, exist_ok=True)

    try:
        wb.save(args.output)
    except Exception as e:
        print(f"ERROR: Could not save output file: {e}")
        sys.exit(1)

    print(f"Output: {args.output}")
    n_sheets = len(required_sheets)
    print(f"Sheets: {n_sheets} | GUI TCs: {gui_tc_count} | FUNC TCs: {func_tc_count} | Common TCs: {common_tc_count}")


if __name__ == "__main__":
    main()
